import os
import PyPDF2
import openpyxl
import re

# Directory path containing PDF files
pdf_directory = 'C:/Users/Admin/Documents/project/nexai_scraper/resume'
excel_file_path = 'extracted_data.xlsx'

# Create or load the Excel workbook and select the active sheet
if os.path.exists(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
else:
    wb = openpyxl.Workbook()

ws = wb.active

# Headers for Excel sheet
headers = ['PDF File', 'Name', 'Contact', 'Email', 'Skills']

# Check if headers already exist, if not, add them
if ws.max_row == 1:
    ws.append(headers)

# Skills to look for in the PDF files
skills = {'html', 'css', 'bootstrap', 'javascript', 'angular', 'typescript', 'git', 'bitbucket', 'agile',
          'jira', 'html5', 'css3', 'aws', 'ec2', 'iam', 's3', 'ses', 'cognito', 'cloudwatch', 'php',
          'python', 'django', 'react', 'nextjs', 'node', 'restful', 'restapi'}

# Regex pattern for contact and email
contact_pattern = re.compile(r'\b(?:\+91)?(?:\d{10}|\d{12})\b')
email_pattern = re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b')

# Loop through each PDF file in the directory
for pdf_filename in os.listdir(pdf_directory):
    if pdf_filename.endswith('.pdf'):
        print("PDF file name:", pdf_filename)

        # Check if the data for this PDF file already exists in the Excel sheet
        existing_data = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row)]
        if pdf_filename in existing_data:
            print(f"Data for {pdf_filename} already exists. Skipping...")
            continue

        # Create a dictionary to store data for the current PDF
        excel_dict = {"PDF File": pdf_filename, "Name": "", "Contact": set(), "Email": set(), "Skills": set()}

        pdf_path = os.path.join(pdf_directory, pdf_filename)

        # Open the PDF file
        with open(pdf_path, 'rb') as pdf_file:
            # Create a PDF reader object
            pdf_reader = PyPDF2.PdfReader(pdf_file)

            # Loop through each page in the PDF
            for page_num in range(len(pdf_reader.pages)):
                # Extract text from the page
                page_obj = pdf_reader.pages[page_num]
                page_text = page_obj.extract_text()

                # Process the extracted text
                for line in page_text.split('\n'):
                    for word in line.split():
                        # Check for contact number
                        contact_match = contact_pattern.match(word)
                        if contact_match:
                            contact_number = contact_match.group()
                            # Add +91 if not present and update the dictionary
                            if '+' not in contact_number:
                                contact_number = f'+91{contact_number}'
                            excel_dict['Contact'].add(contact_number)

                        # Check for email address
                        elif email_pattern.match(word):
                            # Add the email address and handle cases where it's split across two words
                            excel_dict['Email'].add(word)

                        # Check for skills (case-insensitive)
                        elif word.lower() in skills:
                            excel_dict['Skills'].add(word.lower())

        # Remove repeated values from the 'Skills', 'Contact', and 'Email' fields
        excel_dict['Skills'] = ', '.join(sorted(excel_dict['Skills']))
        excel_dict['Contact'] = ', '.join(sorted(excel_dict['Contact']))
        excel_dict['Email'] = ', '.join(sorted(excel_dict['Email']))

        # Add data to the Excel sheet for each PDF file
        ws.append([
            excel_dict['PDF File'],
            excel_dict['Name'],
            excel_dict['Contact'],
            excel_dict['Email'],
            excel_dict['Skills'],  # Skills already processed
        ])

# Save the Excel workbook
wb.save(excel_file_path)

print(f'Data has been extracted and saved to {excel_file_path}')
